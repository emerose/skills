"""Filesystem helpers for archivist: walking an experiment folder, classifying
each file, and extracting a lightweight schema + preview from tabular files.

The schema/preview is what gets embedded for *discovery* ("which file has the
Day-29 QuantiGene numbers?"); the real file is opened separately to read exact
values. Kept dependency-light: CSV/TSV use only the stdlib; XLSX uses ``openpyxl``
when available and degrades gracefully when not. Proprietary binaries (``.pzfx``
GraphPad Prism, ``.eds`` qPCR, ``.cram``/``.vcf`` genomics, …) are catalogued as
descriptors only.
"""

from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Any, Iterator

import _meta

# Never index these — VCS, caches, virtualenvs, the store itself, OS cruft.
IGNORE_DIR_NAMES = {".git", ".archivist", "__pycache__", "node_modules", ".ipynb_checkpoints"}
IGNORE_DIR_SUFFIXES = (".dist-info", ".egg-info")
IGNORE_FILE_NAMES = {".DS_Store", "Thumbs.db"}
# Skip anything inside a virtualenv or site-packages, wherever it sits.
IGNORE_PATH_PARTS = {".venv", "venv", "site-packages"}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def _is_ignored(rel_parts: tuple[str, ...], filename: str) -> bool:
    if filename in IGNORE_FILE_NAMES:
        return True
    if filename.startswith("~$") or filename.startswith("._"):
        return True  # Office lock/temp files and macOS AppleDouble files
    if filename.endswith((".tmp", ".crdownload")):
        return True
    for part in rel_parts:
        if part in IGNORE_DIR_NAMES or part in IGNORE_PATH_PARTS:
            return True
        if any(part.endswith(suf) for suf in IGNORE_DIR_SUFFIXES):
            return True
    return False


def iter_experiment_files(exp_root: Path) -> Iterator[dict[str, Any]]:
    """Yield a descriptor for each indexable file under an experiment folder.

    Each descriptor: ``{abs_path, rel_parts, filename, role, ext, classification}``
    where ``rel_parts`` are the path components below ``exp_root`` (excluding the
    filename), ``role`` per :func:`_meta.role_for_path_parts`, and
    ``classification`` is ``narrative`` | ``tabular`` | ``binary``.
    """
    exp_root = exp_root.resolve()
    for path in sorted(exp_root.rglob("*")):
        if not path.is_file():
            continue
        rel = path.relative_to(exp_root)
        rel_parts = rel.parts[:-1]
        filename = path.name
        if _is_ignored(rel_parts, filename):
            continue
        if path.stat().st_size == 0:
            continue  # empty files carry no content and collapse spuriously by hash
        ext = path.suffix.lower()
        yield {
            "abs_path": path,
            "rel_parts": rel_parts,
            "filename": filename,
            "role": _meta.role_for_path_parts(rel_parts, filename),
            "ext": ext,
            "classification": _meta.classify_ext(ext),
        }


# --------------------------------------------------------------------------- #
# tabular schema / preview
# --------------------------------------------------------------------------- #
def _csv_schema(path: Path, delimiter: str, *, sample_rows: int = 5) -> tuple[dict[str, Any], str]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= sample_rows + 1:
                # read the rest just to count, cheaply
                n_rest = sum(1 for _ in reader)
                total = i + 1 + n_rest
                break
        else:
            total = len(rows)
    header = rows[0] if rows else []
    body = rows[1:]
    schema = {
        "columns": [{"name": c} for c in header],
        "n_rows": max(total - 1, 0),
    }
    preview_lines = [delimiter.join(header)] + [delimiter.join(r) for r in body[:sample_rows]]
    return schema, "\n".join(preview_lines)


def _xlsx_schema(path: Path, *, sample_rows: int = 5) -> tuple[dict[str, Any], str] | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return None
        rows_iter = ws.iter_rows(values_only=True)
        rows = []
        for i, row in enumerate(rows_iter):
            rows.append(row)
            if i >= sample_rows:
                break
        header = [str(c) if c is not None else "" for c in (rows[0] if rows else [])]
        n_rows = (ws.max_row or 1) - 1
        schema = {"columns": [{"name": c} for c in header], "n_rows": max(n_rows, 0),
                  "sheet": ws.title}
        preview_lines = ["\t".join(header)]
        for r in rows[1:sample_rows + 1]:
            preview_lines.append("\t".join("" if c is None else str(c) for c in r))
        return schema, "\n".join(preview_lines)
    finally:
        wb.close()


def schema_and_preview(path: Path) -> tuple[dict[str, Any] | None, str | None]:
    """Best-effort ``(schema, preview)`` for a tabular file.

    Returns ``(None, None)`` for formats we can't crack here (``.pzfx``,
    ``.numbers``, password-protected workbooks, or XLSX when openpyxl is absent) —
    the file is still catalogued, just as a descriptor without column detail.
    """
    ext = path.suffix.lower()
    try:
        if ext in (".csv",):
            return _csv_schema(path, ",")
        if ext in (".tsv",):
            return _csv_schema(path, "\t")
        if ext in (".xlsx", ".xlsm"):
            res = _xlsx_schema(path)
            return res if res else (None, None)
    except Exception:
        # A malformed/locked file shouldn't abort indexing — catalogue it bare.
        return None, None
    return None, None
