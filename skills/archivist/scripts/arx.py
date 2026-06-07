#!/usr/bin/env -S uv run --quiet --script
# /// script
# requires-python = ">=3.12"
# dependencies = ["libkit>=0.2.3", "openpyxl>=3.1", "platformdirs>=4.0", "pyyaml>=6.0"]
# ///
"""arx - a libkit-backed archivist for a tree of scientific experiments.

The managed folder (default: $ARCHIVIST_HOME or the current directory) holds one
subfolder per experiment, named ``K1-YYMMXX - Short Name``, each with the layout
the folder's own LAYOUT.md defines (raw/ data/ protocol/ reports/ analysis/ +
README.md). archivist keeps a libkit store under ``<home>/.archivist/`` (the
single source of truth) that indexes every file for full-text + semantic search
and tracks experiment-level metadata, entities, and cross-references.

libkit IS the store: there is no separate archivist database. Each experiment,
file, and curated-entity note is one libkit document; all archivist fields live
in the document's free-form ``metadata`` JSON.

This CLI is a thin, guardrail layer over the importable modules (`_store`,
`_meta`, `_files`); novel one-off operations can import those directly. Run
`arx <command> --help` for details.
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path
from typing import Any, NoReturn

sys.path.insert(0, str(Path(__file__).resolve().parent))

import _audit  # noqa: E402
import _experiment  # noqa: E402
import _extract  # noqa: E402
import _files  # noqa: E402
import _generate  # noqa: E402
import _intake  # noqa: E402
import _meta  # noqa: E402
import _pr  # noqa: E402
from _store import ArchivistStore, EmbedderConfigError, STORE_DIRNAME  # noqa: E402

# Narrative files larger than this are catalogued as descriptors rather than
# parsed+embedded whole (avoids choking on the multi-hundred-MB raw text dumps).
MAX_EMBED_BYTES = 25 * 1024 * 1024


def _load_dotenv(home: Path) -> None:
    """Load KEY=VALUE pairs from .env files (stdlib only). Real env + earlier
    files win. Search: home, cwd, this script's parents, then ~/.env."""
    here = Path(__file__).resolve()
    candidates = [home / ".env", Path.cwd() / ".env",
                  *[p / ".env" for p in here.parents], Path.home() / ".env"]
    seen: set[Path] = set()
    for env_path in candidates:
        if env_path in seen or not env_path.is_file():
            continue
        seen.add(env_path)
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def die(msg: str, code: int = 1) -> NoReturn:
    print(f"error: {msg}", file=sys.stderr)
    raise SystemExit(code)


def emit_json(obj: Any) -> None:
    import json
    print(json.dumps(obj, ensure_ascii=False, indent=2, default=str))


def _home(args: argparse.Namespace) -> Path:
    return Path(args.home or os.environ.get("ARCHIVIST_HOME") or Path.cwd()).resolve()


def _require_initialized(home: Path) -> None:
    if not (home / STORE_DIRNAME / "catalog.duckdb").exists():
        die(f"no archivist store under {home} — run `arx init --home {home}` first")


def _find_experiment_dir(home: Path, ident: str) -> tuple[Path, dict[str, Any]] | None:
    """Resolve an experiment by exp_id (prefix) or by a path."""
    p = Path(ident)
    if p.is_dir():
        parsed = _meta.parse_experiment_dirname(p.name)
        if parsed:
            return p.resolve(), parsed
    for child in sorted(home.iterdir()):
        if not child.is_dir():
            continue
        parsed = _meta.parse_experiment_dirname(child.name)
        if parsed and (parsed["exp_id"] == ident or child.name == ident):
            return child.resolve(), parsed
    return None


# --------------------------------------------------------------------------- #
# commands
# --------------------------------------------------------------------------- #
async def cmd_init(store: ArchivistStore, args: argparse.Namespace) -> None:
    # opening already created the store; ensure a .gitignore covers it.
    gi = store.home / ".gitignore"
    needed = [f"{STORE_DIRNAME}/", ".DS_Store"]
    existing = gi.read_text(encoding="utf-8").splitlines() if gi.exists() else []
    add = [ln for ln in needed if ln not in existing]
    if add:
        with gi.open("a", encoding="utf-8") as fh:
            if existing and existing[-1].strip():
                fh.write("\n")
            fh.write("\n".join(add) + "\n")
    print(f"initialized archivist store at {store.home / STORE_DIRNAME}")
    if add:
        print(f"  added to .gitignore: {', '.join(add)}")


async def cmd_index(store: ArchivistStore, args: argparse.Namespace) -> None:
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r} under {store.home}")
    exp_dir, parsed = found
    result = await _index_experiment(store, exp_dir, parsed, verbose=not args.json)
    if args.json:
        emit_json(result)
    else:
        print(f"indexed {result['exp_id']}: {result['files_indexed']} files "
              f"({result['narrative']} narrative, {result['tabular']} tabular, "
              f"{result['binary']} binary)")


async def cmd_reindex(store: ArchivistStore, args: argparse.Namespace) -> None:
    results = []
    exp_dirs = [(c, p) for c in sorted(store.home.iterdir())
                if c.is_dir() and (p := _meta.parse_experiment_dirname(c.name))]
    for i, (child, parsed) in enumerate(exp_dirs, 1):
        r = await _index_experiment(store, child.resolve(), parsed, verbose=not args.json)
        results.append(r)
        if not args.json:
            print(f"  [{i}/{len(exp_dirs)}] {r['exp_id']}: {r['files_indexed']} files "
                  f"({r['narrative']}n/{r['tabular']}t/{r['binary']}b)", flush=True)
    if args.json:
        emit_json(results)
    else:
        total = sum(r["files_indexed"] for r in results)
        print(f"indexed {len(results)} experiments, {total} files total")


async def _index_experiment(store: ArchivistStore, exp_dir: Path,
                            parsed: dict[str, Any], *, verbose: bool) -> dict[str, Any]:
    counts = {"narrative": 0, "tabular": 0, "binary": 0, "files_indexed": 0}
    for f in _files.iter_experiment_files(exp_dir):
        abs_path: Path = f["abs_path"]
        rel = store.relpath(abs_path)
        size = abs_path.stat().st_size
        rec: dict[str, Any] = {
            "exp_id": parsed["exp_id"],
            "path": rel,
            "filename": f["filename"],
            "role": f["role"],
            "file_type": f["ext"].lstrip("."),
            "size": size,
            "sha256": _files.sha256_file(abs_path),
        }
        cls = f["classification"]
        try:
            if cls == "narrative" and size <= MAX_EMBED_BYTES:
                try:
                    rec["indexed_as"] = _meta.INDEXED_CONTENT
                    await store.add_file(rec, ingest_path=abs_path)
                except Exception as e:
                    # Parse/loader failure: don't drop the file — catalogue it as a
                    # descriptor so it's still discoverable, and note why.
                    if verbose:
                        print(f"  ! {rel}: {type(e).__name__}: {e} (catalogued as descriptor)",
                              file=sys.stderr)
                    rec["indexed_as"] = _meta.INDEXED_DESCRIPTOR
                    rec["note"] = f"content not embedded ({type(e).__name__}); catalogued only"
                    await store.add_file(rec, card_markdown=_meta.file_card_markdown(rec))
            elif cls == "tabular":
                schema, preview = _files.schema_and_preview(abs_path)
                rec["indexed_as"] = _meta.INDEXED_SCHEMA
                if schema:
                    rec["schema"] = schema
                card = _meta.file_card_markdown(rec, schema=schema, preview=preview)
                await store.add_file(rec, card_markdown=card)
            else:
                rec["indexed_as"] = _meta.INDEXED_DESCRIPTOR
                if cls == "narrative":
                    rec["note"] = "narrative file too large to embed; catalogued only"
                card = _meta.file_card_markdown(rec)
                await store.add_file(rec, card_markdown=card)
        except Exception as e:  # one bad file shouldn't abort the experiment
            if verbose:
                print(f"  ! {rel}: {type(e).__name__}: {e}", file=sys.stderr)
            continue
        counts[cls] += 1
        counts["files_indexed"] += 1

    exp_rec: dict[str, Any] = {
        "exp_id": parsed["exp_id"],
        "name": parsed["name"],
        "title": parsed["name"],
        "folder": store.relpath(exp_dir),
        "file_counts": counts,
    }
    # Structured metadata comes ONLY from the schema'd experiment.yml sidecar — never
    # scraped from README prose. A missing/invalid sidecar leaves metadata minimal
    # (and `check`/`audit` flag it); a bad sidecar surfaces a clear error, not silence.
    sidecar = exp_dir / _experiment.SIDECAR_NAME
    if sidecar.is_file():
        try:
            meta = _experiment.read_sidecar(sidecar)
            meta.pop("exp_id", None)        # folder is authoritative for the id
            exp_rec.update(meta)
            exp_rec.setdefault("name", parsed["name"])
        except _experiment.SidecarError as e:
            print(f"  ! {parsed['exp_id']}: {e}", file=sys.stderr)
            exp_rec["metadata_error"] = str(e)
    await store.upsert_experiment(exp_rec)
    summary = {"exp_id": parsed["exp_id"], **counts}
    if not verbose:
        summary["metadata"] = {k: exp_rec.get(k) for k in
                               ("cro", "cro_study_ids", "assays", "asos", "model", "status")}
    return summary


async def cmd_list(store: ArchivistStore, args: argparse.Namespace) -> None:
    if args.kind == "file":
        recs = await store.files(args.experiment)
    elif args.kind == "entity":
        recs = await store.all_records({"kind": "entity"})
    else:
        recs = await store.experiments()
    recs.sort(key=lambda r: r.get("exp_id") or r.get("path") or r.get("entity_id") or "")
    if args.json:
        emit_json(recs)
        return
    if not recs:
        print("(nothing indexed)")
        return
    for r in recs:
        if args.kind == "file":
            print(f"  [{r.get('role','?'):8}] {r.get('path')}  ({r.get('indexed_as','?')})")
        elif args.kind == "entity":
            print(f"  {r.get('entity_id')}  — {r.get('title') or ''}")
        else:
            fc = r.get("file_counts") or {}
            print(f"  {r.get('exp_id')}  {r.get('name') or r.get('title') or ''}"
                  f"   ({fc.get('files_indexed', 0)} files)")


async def cmd_show(store: ArchivistStore, args: argparse.Namespace) -> None:
    rec = await store.get_experiment(args.experiment)
    if rec is None:
        die(f"no experiment {args.experiment!r} (index it with `arx index`)")
    files = await store.files(args.experiment)
    if args.json:
        emit_json({"experiment": rec, "files": files})
        return
    print(f"{rec.get('exp_id')}: {rec.get('title') or rec.get('name')}")
    for label, key in (("CRO study IDs", "cro_study_ids"), ("CRO", "cro"),
                       ("Status", "status"), ("Folder", "folder")):
        v = rec.get(key)
        if v:
            print(f"  {label}: {', '.join(v) if isinstance(v, list) else v}")
    for label, key in (("Assays", "assays"), ("ASOs", "asos"), ("Model", "model")):
        v = rec.get(key)
        if v:
            print(f"  {label}: {', '.join(v) if isinstance(v, list) else v}")
    print(f"\n  Files ({len(files)}):")
    for line in _generate.files_on_disk_table(files).splitlines():
        print(f"  {line}")


async def cmd_search(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Metadata search across experiments + files (substring over key fields)."""
    needle = args.text.lower()
    hits = []
    for r in await store.all_records():
        hay = " ".join(str(r.get(k, "")) for k in
                       ("exp_id", "name", "title", "cro", "path", "role", "filename")).lower()
        hay += " ".join(str(x) for x in (r.get("cro_study_ids") or []) +
                        (r.get("assays") or []) + (r.get("asos") or []) + (r.get("tags") or [])).lower()
        if needle in hay:
            hits.append(r)
    if args.json:
        emit_json(hits)
        return
    if not hits:
        print("(no matches)")
        return
    for r in hits:
        if r.get("kind") == "experiment":
            print(f"  [exp]  {r.get('exp_id')}  {r.get('title') or r.get('name')}")
        else:
            print(f"  [file] {r.get('exp_id')}  {r.get('path')}")


async def cmd_query(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Semantic + full-text search inside indexed content (libkit hybrid)."""
    filters = {"kind": args.kind} if args.kind else None
    results = await store.query(args.text, limit=args.limit, filters=filters)
    out = []
    for r in results:
        chunk = r.chunk
        meta = chunk.metadata or {}
        out.append({
            "score": r.score,
            "exp_id": meta.get("exp_id"),
            "path": meta.get("path"),
            "kind": meta.get("kind"),
            "text": chunk.text,
        })
    if args.json:
        emit_json(out)
        return
    if not out:
        print("(no results)")
        return
    for h in out:
        loc = h.get("path") or h.get("exp_id") or "?"
        snippet = (h.get("text") or "").strip().replace("\n", " ")[:200]
        print(f"  {loc}\n      {snippet}")


async def cmd_file(store: ArchivistStore, args: argparse.Namespace) -> None:
    rec = await store.get_file(args.path)
    if rec is None:
        die(f"no file record for {args.path!r}")
    emit_json(rec)


async def cmd_read(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Format-aware dump of a tabular file to stdout (for pulling exact values)."""
    path = (store.home / args.path) if not Path(args.path).is_absolute() else Path(args.path)
    if not path.exists():
        die(f"file not found: {path}")
    ext = path.suffix.lower()
    if ext in (".csv", ".tsv", ".xlsx", ".xlsm"):
        if ext in (".csv", ".tsv"):
            print(path.read_text(encoding="utf-8", errors="replace"))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    print(f"# sheet: {ws.title}")
                    for row in ws.iter_rows(values_only=True):
                        print("\t".join("" if c is None else str(c) for c in row))
            finally:
                wb.close()
    else:
        die(f"`read` handles csv/tsv/xlsx; {ext or 'this file'} should be opened directly "
            f"(path: {path})")


async def cmd_entity(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Entities are derived live from experiment records (registry), plus any
    curated notes (kind=entity). `list` aggregates; `show` filters experiments."""
    exps = await store.experiments()
    if args.entity_action == "list":
        agg: dict[str, dict[str, set]] = {"asos": {}, "assays": {}, "cro": {}}
        for e in exps:
            for fld, key in (("asos", "asos"), ("assays", "assays")):
                for v in e.get(key) or []:
                    agg["asos" if fld == "asos" else "assays"].setdefault(v, set()).add(e["exp_id"])
            if e.get("cro"):
                agg["cro"].setdefault(e["cro"], set()).add(e["exp_id"])
        if args.json:
            emit_json({k: {name: sorted(ids) for name, ids in d.items()} for k, d in agg.items()})
            return
        for kind, d in agg.items():
            if d:
                print(f"{kind}:")
                for name, ids in sorted(d.items()):
                    print(f"  {name}  ({len(ids)} experiments)")
    else:  # show
        ident = args.name
        matched = [e["exp_id"] for e in exps
                   if ident in (e.get("asos") or []) or ident in (e.get("assays") or [])
                   or ident == e.get("cro") or ident in (e.get("cro_study_ids") or [])]
        note = await store.get_entity(_meta.entity_slug(ident))
        out = {"entity": ident, "experiments": sorted(matched),
               "curated_note": (note or {}).get("note")}
        emit_json(out) if args.json else print(
            f"{ident}: {len(matched)} experiments\n  " + "\n  ".join(sorted(matched)))


async def cmd_new(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Scaffold a new experiment folder: subdirs + a prose README + a starter
    experiment.yml (the structured metadata), then index it."""
    parsed = _meta.parse_experiment_dirname(f"{args.exp_id} - {args.name}")
    if not parsed:
        die(f"invalid experiment id/name (expected K1-YYMMXX): {args.exp_id!r}")
    folder = store.home / f"{args.exp_id} - {args.name}"
    if folder.exists():
        die(f"folder already exists: {folder}")
    for sub in ("raw", "data", "protocol", "reports", "analysis"):
        (folder / sub).mkdir(parents=True, exist_ok=True)
    (folder / "README.md").write_text(
        _meta.readme_template({"exp_id": args.exp_id, "name": args.name}), encoding="utf-8")
    meta = _experiment.validate({
        "exp_id": args.exp_id, "name": args.name, "status": "planned",
        "cro": args.cro, "model": args.model,
        "cro_study_ids": [args.study_id] if args.study_id else [],
    })
    (folder / _experiment.SIDECAR_NAME).write_text(_experiment.dump_sidecar(meta), encoding="utf-8")
    await _index_experiment(store, folder.resolve(), parsed, verbose=not args.json)
    if args.json:
        emit_json({"created": store.relpath(folder), "exp_id": args.exp_id})
    else:
        print(f"created {store.relpath(folder)} (raw/ data/ protocol/ reports/ analysis/ "
              f"+ README.md + experiment.yml)")
        print(f"indexed as {args.exp_id}; write README.md prose + fill experiment.yml")


async def cmd_intake(store: ArchivistStore, args: argparse.Namespace) -> None:
    """File a delivery (folder or files) into an experiment per LAYOUT.md.

    Copies (never moves) from the source; dry-run by default — review the plan,
    then re-run with --commit to copy + index.
    """
    import shutil

    src = Path(args.source).expanduser()
    if not src.exists():
        die(f"source not found: {src}")
    sources = sorted(p for p in src.rglob("*") if p.is_file()) if src.is_dir() else [src]

    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r} — scaffold it first with `arx new`")
    exp_dir, parsed = found
    plan = _intake.plan_intake(sources, exp_dir)

    if args.json and not args.commit:
        emit_json({"experiment": parsed["exp_id"], "dry_run": True,
                   "plan": [{"src": str(p["src"]), "dest": store.relpath(p["dest"]),
                             "subdir": p["subdir"], "collision": p["exists"]} for p in plan]})
        return
    if not args.commit:
        print(f"intake plan for {parsed['exp_id']} (dry-run — nothing copied):")
        by_sub: dict[str, int] = {}
        for p in plan:
            by_sub[p["subdir"]] = by_sub.get(p["subdir"], 0) + 1
            flag = "  ⚠ overwrites existing" if p["exists"] else ""
            print(f"  {p['subdir']:8} ← {p['src'].name}{flag}")
        print(f"  ({len(plan)} files: " + ", ".join(f"{n} {s}" for s, n in sorted(by_sub.items())) + ")")
        print("re-run with --commit to copy these in and index.")
        return

    copied = 0
    for p in plan:
        p["dest"].parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(p["src"], p["dest"])
        copied += 1
    result = await _index_experiment(store, exp_dir, parsed, verbose=not args.json)
    if args.json:
        emit_json({"experiment": parsed["exp_id"], "copied": copied, "indexed": result})
    else:
        print(f"copied {copied} files into {parsed['exp_id']} and reindexed "
              f"({result['files_indexed']} files total)")


async def cmd_catalog(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Export the experiment catalog: CATALOG.md (human index) + catalog.json."""
    import json

    exps = await store.experiments()
    exps.sort(key=lambda r: r.get("exp_id") or "")
    clean = [{k: v for k, v in e.items() if not k.startswith("_") and k != "content_hash"}
             for e in exps]
    md_path = store.home / "CATALOG.md"
    json_path = store.home / STORE_DIRNAME / "catalog.json"
    md_path.write_text(_meta.catalog_markdown(exps), encoding="utf-8")
    json_path.write_text(json.dumps(clean, ensure_ascii=False, indent=2, default=str, sort_keys=True),
                         encoding="utf-8")
    if args.json:
        emit_json({"experiments": len(exps), "markdown": store.relpath(md_path),
                   "json": store.relpath(json_path)})
    else:
        print(f"wrote {store.relpath(md_path)} and {store.relpath(json_path)} "
              f"({len(exps)} experiments)")


async def _experiment_dirs(store: ArchivistStore, only: str | None):
    """Yield (exp_dir, exp_id) for one experiment or all of them."""
    if only:
        found = _find_experiment_dir(store.home, only)
        if not found:
            die(f"no experiment matching {only!r}")
        yield found[0], found[1]["exp_id"]
        return
    for child in sorted(store.home.iterdir()):
        parsed = _meta.parse_experiment_dirname(child.name) if child.is_dir() else None
        if parsed:
            yield child.resolve(), parsed["exp_id"]


async def cmd_check(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Deterministic structural integrity report (reports only; never mutates)."""
    worklist = []
    async for exp_dir, exp_id in _experiment_dirs(store, args.experiment):
        rec = await store.get_experiment(exp_id) or {"exp_id": exp_id}
        files = await store.files(exp_id)
        flags = _audit.structural_flags(store.home, exp_dir, rec, files)
        if flags:
            worklist.append({"exp_id": exp_id, "flags": flags})
    if args.json:
        emit_json(worklist)
        return
    if not worklist:
        print("✓ no structural issues found")
        return
    for item in worklist:
        print(f"{item['exp_id']}:")
        for f in item["flags"]:
            print(f"    {f}")


async def cmd_audit(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Provenance staleness (experiment.yml fingerprint vs the evidence on disk) +
    a worklist for the parallel-agent semantic pass."""
    report = []
    async for exp_dir, exp_id in _experiment_dirs(store, args.experiment):
        sidecar_path = exp_dir / _experiment.SIDECAR_NAME
        files = await store.files(exp_id)
        entry: dict[str, Any] = {"exp_id": exp_id}
        if not sidecar_path.is_file():
            entry["staleness"] = "no-experiment-yml"
        else:
            try:
                sidecar = _experiment.read_sidecar(sidecar_path)
                st = _experiment.staleness(store.home, exp_dir, sidecar)
                entry["staleness"] = st["state"]
                if st["state"] == "stale":
                    for k in ("changed", "missing", "added", "artifact_changed", "reviewed_at"):
                        if st.get(k):
                            entry[k] = st[k]
            except _experiment.SidecarError as e:
                entry["staleness"] = "invalid-experiment-yml"
                entry["error"] = str(e)
        # source files an agent should read to verify the prose semantically
        entry["source_files"] = [fr["path"] for fr in files
                                 if fr.get("role") in ("data", "report", "raw", "analysis")]
        report.append(entry)
    if args.json:
        emit_json(report)
        return
    for e in report:
        print(f"{e['exp_id']}: {e['staleness']}")
        if e.get("error"):
            print(f"    {e['error']}")
        if e.get("staleness") == "stale":
            if e.get("artifact_changed"):
                print("    README edited since last review")
            for p in e.get("changed", []):
                print(f"    changed: {p}")
            for p in e.get("missing", []):
                print(f"    missing: {p}")
            for p in e.get("added", []):
                print(f"    added (unrecorded): {p}")
            print(f"    last reviewed {e.get('reviewed_at')}")
    print("\nFor the semantic pass, run `arx audit --json` and fan out an agent per "
          "experiment to read its source_files and verify the README prose.")


async def cmd_meta(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Show an experiment's structured metadata (from experiment.yml), or with
    --suggest, print a *draft* sidecar derived heuristically from the README for an
    agent/human to review and write. The tool never writes the sidecar from prose."""
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r}")
    exp_dir, parsed = found
    sidecar = exp_dir / _experiment.SIDECAR_NAME

    if args.suggest:
        readme = exp_dir / "README.md"
        draft: dict[str, Any] = {"exp_id": parsed["exp_id"], "name": parsed["name"]}
        if readme.is_file():
            guess = _extract.extract_from_readme(
                readme.read_text(encoding="utf-8", errors="replace"),
                exp_id=parsed["exp_id"], home=store.home)
            for k in ("title", "cro", "cro_study_ids", "status", "model", "assays", "asos", "related"):
                if guess.get(k):
                    draft[k] = guess[k]
        try:
            draft = _experiment.validate(draft)
        except _experiment.SidecarError:
            pass  # a draft is allowed to be imperfect; the reviewer fixes it
        print(f"# SUGGESTED draft for {sidecar} — REVIEW before saving; not authoritative.")
        print(_experiment.dump_sidecar(draft))
        return

    if not sidecar.is_file():
        die(f"no {_experiment.SIDECAR_NAME} for {parsed['exp_id']} — create one "
            f"(see `arx meta {parsed['exp_id']} --suggest` for a draft)")
    try:
        meta = _experiment.read_sidecar(sidecar)
    except _experiment.SidecarError as e:
        die(str(e))
    emit_json(meta)


async def cmd_fingerprint(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Show the input set `review` would record for an experiment's README right now —
    the in-folder data files (+ any externally-declared inputs), each with its current
    sha256. Lets you see exactly what provenance will track."""
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r}")
    exp_dir, parsed = found
    sidecar = exp_dir / _experiment.SIDECAR_NAME
    declared = []
    if sidecar.is_file():
        meta = _experiment.read_sidecar(sidecar)
        exp_rel = exp_dir.resolve().relative_to(store.home.resolve()).as_posix()
        entry = _experiment._provenance_entry(meta, _experiment.DEFAULT_ARTIFACT) or {}
        declared = [i["path"] for i in (entry.get("inputs") or [])
                    if not i["path"].startswith(exp_rel + "/")]
    inputs, missing = _experiment.resolve_inputs(store.home, exp_dir, declared)
    if args.json:
        emit_json({"exp_id": parsed["exp_id"], "inputs": inputs, "missing": missing})
        return
    for i in inputs:
        print(f"  {i['sha256']}  {i['path']}")
    print(f"({len(inputs)} input files" + (f", {len(missing)} declared-but-missing" if missing else "") + ")")
    for m in missing:
        print(f"  ! missing: {m}", file=sys.stderr)


async def cmd_review(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Mark an experiment's README as verified against its data: record an explicit
    input list (with each file's sha256) + the README's sha + a review date in
    experiment.yml. In-folder data files are included automatically; declare any
    external dependency (e.g. CRO slides under Shared/) with --input <repo-rel path>
    (repeatable; preserved across re-reviews). `audit` then reports per-file drift."""
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r}")
    exp_dir, parsed = found
    sidecar = exp_dir / _experiment.SIDECAR_NAME
    if not sidecar.is_file():
        die(f"no {_experiment.SIDECAR_NAME} for {parsed['exp_id']} — create one first")
    try:
        meta = _experiment.read_sidecar(sidecar)
    except _experiment.SidecarError as e:
        die(str(e))
    updated, missing = _experiment.review(store.home, exp_dir, meta, today=args.date,
                                          extra_inputs=args.input or [])
    sidecar.write_text(_experiment.dump_sidecar(updated), encoding="utf-8")
    entry = _experiment._provenance_entry(updated, _experiment.DEFAULT_ARTIFACT)
    if args.json:
        emit_json({"exp_id": parsed["exp_id"], "provenance": entry, "missing": missing})
    else:
        print(f"stamped {parsed['exp_id']}: README verified against {len(entry['inputs'])} "
              f"input files (reviewed {entry['reviewed_at']})")
        for m in missing:
            print(f"  ! declared input not found on disk: {m}", file=sys.stderr)


async def cmd_pr(store: ArchivistStore, args: argparse.Namespace) -> None:
    """Package working-tree changes into a branch + pull request for review."""
    await _maybe_pr(store, args.paths or None, args.title,
                    args.body or args.title, args, dry_run=args.dry_run)


async def _maybe_pr(store: ArchivistStore, paths, title: str, body: str,
                    args: argparse.Namespace, *, dry_run: bool = False) -> None:
    try:
        result = _pr.create_pr(store.home, title=title, body=body,
                               paths=paths or _changed_paths(store.home),
                               dry_run=dry_run)
    except _pr.GitError as e:
        die(str(e))
    if args.json:
        emit_json(result)
    elif result.get("pr_url"):
        print(f"opened PR: {result['pr_url']}")
    elif result.get("dry_run"):
        print("dry-run — would run:\n  " + "\n  ".join(result["steps"]))
    else:
        print(f"committed to branch {result['branch']} (not pushed)")


def _changed_paths(home: Path) -> list[str]:
    import subprocess
    # -z gives NUL-separated, UNquoted paths (paths here contain spaces), so we can
    # hand them straight to `git add --` without quoting/escaping surprises.
    out = subprocess.run(["git", "-C", str(home), "status", "--porcelain", "-z"],
                         capture_output=True, text=True).stdout
    return [entry[3:] for entry in out.split("\0") if entry.strip()]


# --------------------------------------------------------------------------- #
# dispatch
# --------------------------------------------------------------------------- #
COMMANDS = {
    "init": cmd_init,
    "index": cmd_index,
    "reindex": cmd_reindex,
    "list": cmd_list,
    "show": cmd_show,
    "search": cmd_search,
    "query": cmd_query,
    "file": cmd_file,
    "read": cmd_read,
    "entity": cmd_entity,
    "new": cmd_new,
    "intake": cmd_intake,
    "catalog": cmd_catalog,
    "check": cmd_check,
    "audit": cmd_audit,
    "meta": cmd_meta,
    "fingerprint": cmd_fingerprint,
    "review": cmd_review,
    "pr": cmd_pr,
}


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="arx", description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--home", help="managed data folder (default: $ARCHIVIST_HOME or cwd)")
    sub = p.add_subparsers(dest="command", required=True)

    def add(name: str, help_: str) -> argparse.ArgumentParser:
        sp = sub.add_parser(name, help=help_)
        sp.add_argument("--json", action="store_true", help="machine-readable output")
        return sp

    add("init", "create the libkit store and .gitignore under the data folder")
    sp = add("index", "index one experiment folder (by exp_id or path)")
    sp.add_argument("experiment")
    add("reindex", "index every experiment folder under the data folder")
    sp = add("list", "list experiments (default), files, or entities")
    sp.add_argument("--kind", choices=["experiment", "file", "entity"], default="experiment")
    sp.add_argument("--experiment", help="when --kind file: limit to this exp_id")
    sp = add("show", "show one experiment and its files")
    sp.add_argument("experiment")
    sp = add("search", "metadata search across experiments and files")
    sp.add_argument("text")
    sp = add("query", "semantic + full-text search inside indexed content")
    sp.add_argument("text")
    sp.add_argument("--limit", type=int, default=8)
    sp.add_argument("--kind", choices=["experiment", "file", "entity"], default=None)
    sp = add("file", "show one file record (by relative path)")
    sp.add_argument("path")
    sp = add("read", "dump a tabular file (csv/tsv/xlsx) to stdout")
    sp.add_argument("path")
    sp = add("entity", "list derived entities or show one entity's experiments")
    sp.add_argument("entity_action", choices=["list", "show"])
    sp.add_argument("name", nargs="?", help="entity name (for show)")
    sp = add("new", "scaffold a new experiment folder (subdirs + README template)")
    sp.add_argument("exp_id", help="internal id, e.g. K1-000003")
    sp.add_argument("name", help="short name, e.g. 'Rat IT Chronic Tox'")
    sp.add_argument("--cro", help="contract research org")
    sp.add_argument("--study-id", help="external/CRO study id")
    sp.add_argument("--model", help="species/model")
    sp = add("intake", "file a delivery (folder/files) into an experiment per LAYOUT.md")
    sp.add_argument("experiment", help="target experiment (exp_id or folder)")
    sp.add_argument("source", help="a delivery folder or a single file (copied, not moved)")
    sp.add_argument("--commit", action="store_true", help="actually copy + index (default: dry-run)")
    add("catalog", "export the experiment catalog (CATALOG.md + catalog.json)")
    sp = add("check", "structural integrity report (missing/unindexed files, layout, redundant archives)")
    sp.add_argument("experiment", nargs="?", help="limit to one experiment (default: all)")
    sp = add("audit", "staleness of generated docs + a worklist for the semantic pass")
    sp.add_argument("experiment", nargs="?", help="limit to one experiment (default: all)")
    sp = add("meta", "show an experiment's structured metadata (experiment.yml)")
    sp.add_argument("experiment")
    sp.add_argument("--suggest", action="store_true",
                    help="print a heuristic draft sidecar from the README to review (not authoritative)")
    sp = add("fingerprint", "show the input files (+ sha256) review would record for an experiment")
    sp.add_argument("experiment")
    sp = add("review", "record provenance after verifying the README vs its data (explicit input list)")
    sp.add_argument("experiment")
    sp.add_argument("--input", action="append", metavar="REPO_REL_PATH",
                    help="declare an external dependency (repeatable; e.g. a Shared/ CRO file)")
    sp.add_argument("--date", help="review date YYYY-MM-DD (default: today)")
    sp = add("pr", "package working-tree changes into a branch + pull request")
    sp.add_argument("title")
    sp.add_argument("paths", nargs="*", help="paths to include (default: all changes)")
    sp.add_argument("--body", help="PR body")
    sp.add_argument("--dry-run", action="store_true", help="show the git/gh steps, do nothing")
    return p


class _HomeOnly:
    """Lightweight stand-in for commands that need only the data folder, not the
    libkit store (so `pr` doesn't require an embedding backend)."""

    def __init__(self, home: Path) -> None:
        self.home = home

    def relpath(self, path: Path) -> str:
        try:
            return str(path.resolve().relative_to(self.home.resolve()))
        except ValueError:
            return str(path)


async def _run(args: argparse.Namespace) -> None:
    home = _home(args)
    _load_dotenv(home)
    handler = COMMANDS[args.command]
    if args.command == "init":
        home.mkdir(parents=True, exist_ok=True)
    else:
        _require_initialized(home)
    if args.command == "pr":            # pure git; no libkit store needed
        await handler(_HomeOnly(home), args)  # type: ignore[arg-type]
        return
    try:
        store = ArchivistStore.open(home)
    except EmbedderConfigError as e:
        die(str(e))
    try:
        await handler(store, args)
    finally:
        await store.close()


def main() -> None:
    args = build_parser().parse_args()
    try:
        asyncio.run(_run(args))
    except KeyboardInterrupt:
        die("interrupted", code=130)


if __name__ == "__main__":
    main()
