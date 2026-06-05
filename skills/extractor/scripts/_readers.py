"""Shared, deterministic readers for CRO source files → tidy rows.

Each reader returns (header: list[str], rows: list[list]) so a recipe can write a
CSV and the extractor can record provenance. Readers are pure functions of the
source bytes — re-running on the same input yields identical output (diffable).

Formats: .xlsx (openpyxl) and GraphPad Prism .pzfx (XML) cover the large majority
of measurement data. Add new readers here so recipes stay thin.
"""
from __future__ import annotations
import datetime as _dt
import xml.etree.ElementTree as ET
from pathlib import Path


def _fmt_dt(d) -> str:
    """A date/datetime cell → faithful ISO string. A datetime whose time part is
    midnight is a *date* cell (Excel stores dates as datetimes at 00:00:00) and is
    rendered date-only ('2023-05-19'); a real time component is kept
    ('2023-05-19 14:30:00'). Matches how dates are normally written by hand, so a
    faithful extraction is byte-comparable to hand-curated date columns."""
    if isinstance(d, _dt.datetime):
        if (d.hour, d.minute, d.second, d.microsecond) == (0, 0, 0, 0):
            return d.date().isoformat()
        return d.isoformat(sep=" ")
    return d.isoformat()  # datetime.date / datetime.time


# --------------------------------------------------------------------------- #
# xlsx
# --------------------------------------------------------------------------- #
def _fmt(v) -> str:
    """Stable cell → string. Integers stay integers; floats keep their value;
    date/datetime cells render as clean ISO dates; None/blank → ''. (Faithful to
    the cell value, no rounding.)"""
    if v is None:
        return ""
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return _fmt_dt(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)

def _read_xlsx_rows(path: Path, sheet: str | None) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        wb.close()
        raise ValueError(f"no worksheet {sheet!r} in {path}")
    rows = [[_fmt(c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_xls_rows(path: Path, sheet: str | None) -> list[list[str]]:
    """Legacy .xls via xlrd (openpyxl can't read .xls)."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)

    def cell(c) -> str:
        if c.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return ""
        if c.ctype == xlrd.XL_CELL_NUMBER:
            return str(int(c.value)) if float(c.value).is_integer() else str(c.value)
        if c.ctype == xlrd.XL_CELL_DATE:
            return _fmt_dt(xlrd.xldate.xldate_as_datetime(c.value, wb.datemode))
        return str(c.value)

    return [[cell(ws.cell(r, col)) for col in range(ws.ncols)] for r in range(ws.nrows)]


def read_xlsx_sheet(path: Path, sheet: str | None = None,
                    drop_blank_rows: bool = True) -> tuple[list[str], list[list[str]]]:
    """Dump one spreadsheet worksheet faithfully (`.xlsx` via openpyxl, `.xls` via
    xlrd): row 1 is the header, the rest are data. `drop_blank_rows` removes rows
    where every cell is blank (trailing padding / spacer rows) — it never drops a
    row that carries any value."""
    suffix = Path(path).suffix.lower()
    rows = _read_xls_rows(path, sheet) if suffix == ".xls" else _read_xlsx_rows(path, sheet)
    if not rows:
        return [], []
    header, data = rows[0], rows[1:]
    if drop_blank_rows:
        data = [r for r in data if any(c != "" for c in r)]
    return header, data

# --------------------------------------------------------------------------- #
# GraphPad Prism .pzfx
# --------------------------------------------------------------------------- #
def _tag(t: str) -> str:
    return t.split("}")[-1]

def _direct_title(col: ET.Element) -> str:
    """The column's own <Title> (direct child), not a nested one."""
    for ch in col:
        if _tag(ch.tag) == "Title":
            return "".join(ch.itertext()).strip()
    return ""

def _subcolumns(col: ET.Element) -> list[list[str]]:
    out = []
    for sc in col:
        if _tag(sc.tag) == "Subcolumn":
            out.append([(d.text or "").strip() for d in sc if _tag(d.tag) == "d"])
    return out

def read_pzfx_structured(path: Path):
    """Structured view of each table: (table_title, x_values, [(y_title, [subcolumns])]).
    `x_values` is the first X-family subcolumn; each Y column keeps its subcolumns
    (replicates) intact — the basis for a tidy long-format reshape."""
    root = ET.parse(path).getroot()
    out = []
    for t in [el for el in root.iter() if _tag(el.tag) == "Table"]:
        xvals: list[str] = []
        ycols: list[tuple[str, list[list[str]]]] = []
        for col in t:
            k = _tag(col.tag)
            if k in ("XColumn", "XAdvancedColumn"):
                if not xvals:
                    subs = _subcolumns(col)
                    if subs:
                        xvals = subs[0]
            elif k == "YColumn":
                ycols.append((_direct_title(col), _subcolumns(col)))
        out.append((_direct_title(t), xvals, ycols))
    return out


def read_pzfx(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    """Return one (table_title, header, rows) per data table, flattened wide:
    leading '' index column (0..n-1), then each X-family subcolumn as `_<k>`,
    then each Y column's subcolumns as `<col title>_<g>` where g is a running
    index across all Y subcolumns in the table. Mirrors the conventional
    pandas-style pzfx dump so the layout is recognizable and value-faithful."""
    root = ET.parse(path).getroot()
    tables = [el for el in root.iter() if _tag(el.tag) == "Table"]
    results = []
    for t in tables:
        title = _direct_title(t)
        xcols, ycols = [], []
        for col in t:
            k = _tag(col.tag)
            if k in ("XColumn", "XAdvancedColumn"):
                xcols.append(col)
            elif k == "YColumn":
                ycols.append(col)
        cols: list[tuple[str, list[str]]] = []          # (name, values)
        for xc in xcols:
            for sub in _subcolumns(xc):
                cols.append((f"_{len(cols)}", sub))
        g = 0
        for yc in ycols:
            ytitle = _direct_title(yc)
            for sub in _subcolumns(yc):
                cols.append((f"{ytitle}_{g}", sub))
                g += 1
        nrows = max((len(v) for _, v in cols), default=0)
        header = [""] + [name for name, _ in cols]
        rows = []
        for i in range(nrows):
            row = [str(i)]
            for _, vals in cols:
                row.append(vals[i] if i < len(vals) else "")
            rows.append(row)
        results.append((title, header, rows))
    return results


# --------------------------------------------------------------------------- #
# Word .docx (CRO study reports whose tables are the only machine-readable source)
# --------------------------------------------------------------------------- #
def read_docx_tables(path: Path) -> list[list[list[str]]]:
    """Return every table in a .docx as a list of tables, each a list of rows, each
    row a list of cell strings (text only; whitespace collapsed). Deterministic.

    Some CRO deliverables ship only as a Word report (no spreadsheet); its tables are
    then the raw source. Tables are returned in document order so a recipe can select
    by index. Merged cells repeat their text across the spanned grid positions (the
    python-docx default), which keeps every row the table's full width."""
    import docx  # provided via the engine's deps
    doc = docx.Document(str(path))
    out: list[list[list[str]]] = []
    for t in doc.tables:
        rows = []
        for row in t.rows:
            rows.append([" ".join(str(c.text).split()) for c in row.cells])
        out.append(rows)
    return out
