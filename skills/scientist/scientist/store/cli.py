"""Command handlers + argparse wiring for the store subcommands (the former ``arx``
CLI), folded into the ``sci`` entry point.

This is a thin, guardrail layer over the importable modules (:mod:`_store`,
:mod:`_meta`, :mod:`_files`, …) and the shared :mod:`provenance` core. All
``experiment.yml`` access (read/validate/write the sidecar, record provenance,
staleness, review inputs) routes through :mod:`provenance` — never re-implemented.

Configuration: the data-tree root resolves from ``--home``, else ``$SCIENTIST_HOME``,
else cwd; the store dir is ``.scientist/catalog.duckdb``.
Third-party keys (``DEEPINFRA_API_KEY`` / ``DATALAB_API_KEY``) are untouched.

``register(subparsers)`` adds the store subcommands to an existing ``sci`` parser;
``dispatch(args)`` runs the selected one (sync wrapper over the async handler).
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path
from typing import Any, NoReturn

from .. import provenance

from . import _audit, _files, _generate, _intake, _meta, _pr
from ._store import STORE_DIRNAME, Store, EmbedderConfigError

# Narrative files larger than this are catalogued as descriptors rather than
# parsed+embedded whole (avoids choking on the multi-hundred-MB raw text dumps).
MAX_EMBED_BYTES = 25 * 1024 * 1024

# The set of subcommands this module owns (so sci can route them here).
STORE_COMMANDS = (
    "init", "index", "reindex", "index-claims", "list", "show", "search", "query",
    "file", "read", "entity", "new", "intake", "meta", "review", "fingerprint",
    "catalog", "check", "audit", "pr",
)


def _load_dotenv(home: Path) -> None:
    """Load KEY=VALUE pairs from .env files (stdlib only). Real env + earlier
    files win. Search: home, cwd, this script's parents, then ~/.env."""
    here = Path(__file__).resolve()
    candidates = [home / ".env", Path.cwd() / ".env",
                  *[p / ".env" for p in here.parents], Path.home() / ".env"]
    seen: set[Path] = set()
    for env_path in candidates:
        if env_path in seen or not env_path.is_file():
            continue
        seen.add(env_path)
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def die(msg: str, code: int = 1) -> NoReturn:
    print(f"error: {msg}", file=sys.stderr)
    raise SystemExit(code)


def emit_json(obj: Any) -> None:
    import json
    print(json.dumps(obj, ensure_ascii=False, indent=2, default=str))


def _home(args: argparse.Namespace) -> Path:
    return Path(args.home or os.environ.get("SCIENTIST_HOME")
                or Path.cwd()).resolve()


def _require_initialized(home: Path) -> None:
    if not (home / STORE_DIRNAME / "catalog.duckdb").exists():
        die(f"no scientist store under {home} — run `sci init --home {home}` first")


def _find_experiment_dir(home: Path, ident: str) -> tuple[Path, dict[str, Any]] | None:
    """Resolve an experiment by exp_id (prefix) or by a path."""
    p = Path(ident)
    if p.is_dir():
        parsed = _meta.parse_experiment_dirname(p.name)
        if parsed:
            return p.resolve(), parsed
    for child in sorted(home.iterdir()):
        if not child.is_dir():
            continue
        parsed = _meta.parse_experiment_dirname(child.name)
        if parsed and (parsed["exp_id"] == ident or child.name == ident):
            return child.resolve(), parsed
    return None


# --------------------------------------------------------------------------- #
# commands
# --------------------------------------------------------------------------- #
async def cmd_init(store: Store, _args: argparse.Namespace) -> None:
    # opening already created the store; ensure a .gitignore covers it.
    gi = store.home / ".gitignore"
    needed = [f"{STORE_DIRNAME}/", ".DS_Store"]
    existing = gi.read_text(encoding="utf-8").splitlines() if gi.exists() else []
    add = [ln for ln in needed if ln not in existing]
    if add:
        with gi.open("a", encoding="utf-8") as fh:
            if existing and existing[-1].strip():
                fh.write("\n")
            fh.write("\n".join(add) + "\n")
    print(f"initialized scientist store at {store.home / store._store_dirname}")
    if add:
        print(f"  added to .gitignore: {', '.join(add)}")


async def cmd_index(store: Store, args: argparse.Namespace) -> None:
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r} under {store.home}")
    exp_dir, parsed = found
    result = await _index_experiment(store, exp_dir, parsed, verbose=not args.json)
    if args.json:
        emit_json(result)
    else:
        print(f"indexed {result['exp_id']}: {result['files_indexed']} files "
              f"({result['narrative']} narrative, {result['tabular']} tabular, "
              f"{result['binary']} binary)")


async def cmd_reindex(store: Store, args: argparse.Namespace) -> None:
    results = []
    exp_dirs = [(c, p) for c in sorted(store.home.iterdir())
                if c.is_dir() and (p := _meta.parse_experiment_dirname(c.name))]
    for i, (child, parsed) in enumerate(exp_dirs, 1):
        r = await _index_experiment(store, child.resolve(), parsed, verbose=not args.json)
        results.append(r)
        if not args.json:
            print(f"  [{i}/{len(exp_dirs)}] {r['exp_id']}: {r['files_indexed']} files "
                  f"({r['narrative']}n/{r['tabular']}t/{r['binary']}b)", flush=True)
    if args.json:
        emit_json(results)
    else:
        total = sum(r["files_indexed"] for r in results)
        print(f"indexed {len(results)} experiments, {total} files total")


async def _index_experiment(store: Store, exp_dir: Path,
                            parsed: dict[str, Any], *, verbose: bool) -> dict[str, Any]:
    counts = {"narrative": 0, "tabular": 0, "binary": 0, "files_indexed": 0}
    for f in _files.iter_experiment_files(exp_dir):
        abs_path: Path = f["abs_path"]
        rel = store.relpath(abs_path)
        size = abs_path.stat().st_size
        rec: dict[str, Any] = {
            "exp_id": parsed["exp_id"],
            "path": rel,
            "filename": f["filename"],
            "role": f["role"],
            "file_type": f["ext"].lstrip("."),
            "size": size,
            "sha256": _files.sha256_file(abs_path),
        }
        cls = f["classification"]
        try:
            if cls == "narrative" and size <= MAX_EMBED_BYTES:
                try:
                    rec["indexed_as"] = _meta.INDEXED_CONTENT
                    await store.add_file(rec, ingest_path=abs_path)
                except Exception as e:
                    # Parse/loader failure: don't drop the file — catalogue it as a
                    # descriptor so it's still discoverable, and note why.
                    if verbose:
                        print(f"  ! {rel}: {type(e).__name__}: {e} (catalogued as descriptor)",
                              file=sys.stderr)
                    rec["indexed_as"] = _meta.INDEXED_DESCRIPTOR
                    rec["note"] = f"content not embedded ({type(e).__name__}); catalogued only"
                    await store.add_file(rec, card_markdown=_meta.file_card_markdown(rec))
            elif cls == "tabular":
                schema, preview = _files.schema_and_preview(abs_path)
                rec["indexed_as"] = _meta.INDEXED_SCHEMA
                if schema:
                    rec["schema"] = schema
                card = _meta.file_card_markdown(rec, schema=schema, preview=preview)
                await store.add_file(rec, card_markdown=card)
            else:
                rec["indexed_as"] = _meta.INDEXED_DESCRIPTOR
                if cls == "narrative":
                    rec["note"] = "narrative file too large to embed; catalogued only"
                card = _meta.file_card_markdown(rec)
                await store.add_file(rec, card_markdown=card)
        except Exception as e:  # one bad file shouldn't abort the experiment
            if verbose:
                print(f"  ! {rel}: {type(e).__name__}: {e}", file=sys.stderr)
            continue
        counts[cls] += 1
        counts["files_indexed"] += 1

    exp_rec: dict[str, Any] = {
        "exp_id": parsed["exp_id"],
        "name": parsed["name"],
        "title": parsed["name"],
        "folder": store.relpath(exp_dir),
        "file_counts": counts,
    }
    # Structured metadata comes ONLY from the schema'd experiment.yml sidecar — never
    # scraped from README prose. A missing/invalid sidecar leaves metadata minimal
    # (and `check`/`audit` flag it); a bad sidecar surfaces a clear error, not silence.
    sidecar = exp_dir / provenance.SIDECAR_NAME
    if sidecar.is_file():
        try:
            meta = provenance.read_sidecar(exp_dir)
            meta.pop("exp_id", None)        # folder is authoritative for the id
            meta.pop("provenance", None)    # the ledger isn't experiment-card metadata
            exp_rec.update(meta)
            exp_rec.setdefault("name", parsed["name"])
        except provenance.SidecarError as e:
            print(f"  ! {parsed['exp_id']}: {e}", file=sys.stderr)
            exp_rec["metadata_error"] = str(e)
    await store.upsert_experiment(exp_rec)
    summary = {"exp_id": parsed["exp_id"], **counts}
    if not verbose:
        summary["metadata"] = {k: exp_rec.get(k) for k in
                               ("cro", "cro_study_ids", "assays", "asos", "model", "status")}
    return summary


def _load_grounding_report(exp_dir: Path, override: str | None) -> tuple[Path, list[dict[str, Any]]]:
    """Locate + parse the grounding_report.json for an experiment.

    Search order: ``--report PATH`` if given, else ``<exp>/analysis/grounding_report.json``
    then ``<exp>/grounding_report.json``. Returns the resolved path + the claims list.
    Dies with a clear, actionable error if no report is found or it's malformed.
    """
    import json

    if override:
        candidates = [Path(override)]
    else:
        candidates = [exp_dir / "analysis" / "grounding_report.json",
                      exp_dir / "grounding_report.json"]
    report_path = next((p for p in candidates if p.is_file()), None)
    if report_path is None:
        looked = ", ".join(str(p) for p in candidates)
        die(f"no grounding report found (looked: {looked}). Run the claims first, e.g.\n"
            f"  uv run --with-editable skills/scientist pytest "
            f"\"{exp_dir / 'analysis' / 'claims'}\"")
    try:
        data = json.loads(report_path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as e:
        die(f"could not read grounding report {report_path}: {e}")
    claims = data.get("claims") if isinstance(data, dict) else data
    if not isinstance(claims, list):
        die(f"grounding report {report_path} has no 'claims' list")
    return report_path, claims


async def cmd_index_claims(store: Store, args: argparse.Namespace) -> None:
    """Index the grounded claims from an experiment's grounding_report.json into the
    libkit store as ``kind=claim`` documents, then prune any claims that have been
    removed from the report (rebuildable store)."""
    import json

    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r} under {store.home}")
    exp_dir, parsed = found
    exp_id = parsed["exp_id"]
    report_path, claims = _load_grounding_report(exp_dir, args.report)

    indexed_ids: list[str] = []
    for claim in claims:
        nodeid = claim.get("id") or ""
        claim_id = _meta.claim_id_for(exp_id, nodeid)
        rec: dict[str, Any] = {
            "exp_id": exp_id,
            "claim_id": claim_id,
            "statement": claim.get("statement") or "",
            "outcome": claim.get("outcome"),
            "strength": claim.get("strength"),
            "claim_kind": claim.get("kind"),
            "caveats": claim.get("caveats"),
            "evidence_json": json.dumps(claim.get("evidence") or {},
                                        ensure_ascii=False, sort_keys=True, default=str),
            "inputs": [{"path": i.get("path"), "sha256": i.get("sha256")}
                       for i in (claim.get("inputs") or [])],
            "source": nodeid,
        }
        await store.upsert_claim(rec)
        indexed_ids.append(claim_id)

    pruned = await store.replace_experiment_claims(exp_id, indexed_ids)
    if args.json:
        emit_json({"exp_id": exp_id, "report": store.relpath(report_path),
                   "indexed": len(indexed_ids), "pruned": pruned})
    else:
        print(f"indexed {len(indexed_ids)} claims for {exp_id} "
              f"(from {store.relpath(report_path)}); pruned {pruned} stale")


async def cmd_list(store: Store, args: argparse.Namespace) -> None:
    if args.kind == "file":
        recs = await store.files(args.experiment)
    elif args.kind == "entity":
        recs = await store.all_records({"kind": "entity"})
    elif args.kind == "claim":
        recs = await store.claims(args.experiment)
    elif args.kind == "report":
        recs = await store.reports(args.experiment)
    else:
        recs = await store.experiments()
    recs.sort(key=lambda r: r.get("exp_id") or r.get("path")
              or r.get("entity_id") or r.get("claim_id") or r.get("report_id") or "")
    if args.json:
        emit_json(recs)
        return
    if not recs:
        print("(nothing indexed)")
        return
    for r in recs:
        if args.kind == "file":
            print(f"  [{r.get('role','?'):8}] {r.get('path')}  ({r.get('indexed_as','?')})")
        elif args.kind == "entity":
            print(f"  {r.get('entity_id')}  — {r.get('title') or ''}")
        elif args.kind == "claim":
            outcome = r.get("outcome") or "?"
            label = _meta.CLAIM_OUTCOME_LABEL.get(outcome, outcome)
            stmt = (r.get("statement") or "").strip().replace("\n", " ")[:90]
            print(f"  [{label} · {r.get('strength','?')}] {r.get('exp_id')}  {stmt}")
        elif args.kind == "report":
            title = (r.get("title") or r.get("slug") or "").strip().replace("\n", " ")[:80]
            print(f"  [{r.get('scope','?')}] {r.get('report_id')}  {title}")
        else:
            fc = r.get("file_counts") or {}
            print(f"  {r.get('exp_id')}  {r.get('name') or r.get('title') or ''}"
                  f"   ({fc.get('files_indexed', 0)} files)")


async def cmd_show(store: Store, args: argparse.Namespace) -> None:
    rec = await store.get_experiment(args.experiment)
    if rec is None:
        die(f"no experiment {args.experiment!r} (index it with `sci index`)")
    files = await store.files(args.experiment)
    if args.json:
        emit_json({"experiment": rec, "files": files})
        return
    print(f"{rec.get('exp_id')}: {rec.get('title') or rec.get('name')}")
    for label, key in (("CRO study IDs", "cro_study_ids"), ("CRO", "cro"),
                       ("Status", "status"), ("Folder", "folder")):
        v = rec.get(key)
        if v:
            print(f"  {label}: {', '.join(v) if isinstance(v, list) else v}")
    for label, key in (("Assays", "assays"), ("ASOs", "asos"), ("Model", "model")):
        v = rec.get(key)
        if v:
            print(f"  {label}: {', '.join(v) if isinstance(v, list) else v}")
    print(f"\n  Files ({len(files)}):")
    for line in _generate.files_on_disk_table(files).splitlines():
        print(f"  {line}")


async def cmd_search(store: Store, args: argparse.Namespace) -> None:
    """Metadata search across experiments + files (substring over key fields)."""
    needle = args.text.lower()
    hits = []
    for r in await store.all_records():
        hay = " ".join(str(r.get(k, "")) for k in
                       ("exp_id", "name", "title", "cro", "path", "role", "filename")).lower()
        hay += " ".join(str(x) for x in (r.get("cro_study_ids") or []) +
                        (r.get("assays") or []) + (r.get("asos") or []) + (r.get("tags") or [])).lower()
        if needle in hay:
            hits.append(r)
    if args.json:
        emit_json(hits)
        return
    if not hits:
        print("(no matches)")
        return
    for r in hits:
        if r.get("kind") == "experiment":
            print(f"  [exp]  {r.get('exp_id')}  {r.get('title') or r.get('name')}")
        else:
            print(f"  [file] {r.get('exp_id')}  {r.get('path')}")


async def cmd_query(store: Store, args: argparse.Namespace) -> None:
    """Semantic + full-text search inside indexed content (libkit hybrid)."""
    filters = {"kind": args.kind} if args.kind else None
    results = await store.query(args.text, limit=args.limit, filters=filters)
    out = []
    for r in results:
        chunk = r.chunk
        meta = chunk.metadata or {}
        hit = {
            "score": r.score,
            "exp_id": meta.get("exp_id"),
            "path": meta.get("path"),
            "kind": meta.get("kind"),
            "text": chunk.text,
        }
        if meta.get("kind") == "claim":
            # Surface the judgment so a contradicted/weak claim is never shown as
            # plain positive evidence.
            hit["outcome"] = meta.get("outcome")
            hit["strength"] = meta.get("strength")
            hit["claim_kind"] = meta.get("claim_kind")
            hit["statement"] = meta.get("statement")
            hit["claim_id"] = meta.get("claim_id")
        elif meta.get("kind") == "report":
            hit["report_id"] = meta.get("report_id")
            hit["scope"] = meta.get("scope")
            hit["report_title"] = meta.get("title")
        out.append(hit)
    if args.json:
        emit_json(out)
        return
    if not out:
        print("(no results)")
        return
    for h in out:
        if h.get("kind") == "claim":
            outcome = h.get("outcome") or "?"
            label = _meta.CLAIM_OUTCOME_LABEL.get(outcome, outcome)
            stmt = (h.get("statement") or h.get("text") or "").strip().replace("\n", " ")[:200]
            print(f"  [claim · {label} · strength: {h.get('strength','?')}] {h.get('exp_id')}\n"
                  f"      {stmt}")
            continue
        if h.get("kind") == "report":
            snippet = (h.get("text") or "").strip().replace("\n", " ")[:200]
            print(f"  [report · {h.get('scope','?')}] {h.get('report_id')}"
                  f"  {h.get('report_title') or ''}\n      {snippet}")
            continue
        loc = h.get("path") or h.get("exp_id") or "?"
        snippet = (h.get("text") or "").strip().replace("\n", " ")[:200]
        print(f"  {loc}\n      {snippet}")


async def cmd_file(store: Store, args: argparse.Namespace) -> None:
    rec = await store.get_file(args.path)
    if rec is None:
        die(f"no file record for {args.path!r}")
    emit_json(rec)


async def cmd_read(store: Store, args: argparse.Namespace) -> None:
    """Format-aware dump of a tabular file to stdout (for pulling exact values)."""
    path = (store.home / args.path) if not Path(args.path).is_absolute() else Path(args.path)
    if not path.exists():
        die(f"file not found: {path}")
    ext = path.suffix.lower()
    if ext in (".csv", ".tsv", ".xlsx", ".xlsm"):
        if ext in (".csv", ".tsv"):
            print(path.read_text(encoding="utf-8", errors="replace"))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    print(f"# sheet: {ws.title}")
                    for row in ws.iter_rows(values_only=True):
                        print("\t".join("" if c is None else str(c) for c in row))
            finally:
                wb.close()
    else:
        die(f"`read` handles csv/tsv/xlsx; {ext or 'this file'} should be opened directly "
            f"(path: {path})")


async def cmd_entity(store: Store, args: argparse.Namespace) -> None:
    """Entities are derived live from experiment records (registry), plus any
    curated notes (kind=entity). `list` aggregates; `show` filters experiments."""
    exps = await store.experiments()
    if args.entity_action == "list":
        agg: dict[str, dict[str, set]] = {"asos": {}, "assays": {}, "cro": {}}
        for e in exps:
            for fld, key in (("asos", "asos"), ("assays", "assays")):
                for v in e.get(key) or []:
                    agg["asos" if fld == "asos" else "assays"].setdefault(v, set()).add(e["exp_id"])
            if e.get("cro"):
                agg["cro"].setdefault(e["cro"], set()).add(e["exp_id"])
        if args.json:
            emit_json({k: {name: sorted(ids) for name, ids in d.items()} for k, d in agg.items()})
            return
        for kind, d in agg.items():
            if d:
                print(f"{kind}:")
                for name, ids in sorted(d.items()):
                    print(f"  {name}  ({len(ids)} experiments)")
    else:  # show
        ident = args.name
        matched = [e["exp_id"] for e in exps
                   if ident in (e.get("asos") or []) or ident in (e.get("assays") or [])
                   or ident == e.get("cro") or ident in (e.get("cro_study_ids") or [])]
        note = await store.get_entity(_meta.entity_slug(ident))
        out = {"entity": ident, "experiments": sorted(matched),
               "curated_note": (note or {}).get("note")}
        emit_json(out) if args.json else print(
            f"{ident}: {len(matched)} experiments\n  " + "\n  ".join(sorted(matched)))


async def cmd_new(store: Store, args: argparse.Namespace) -> None:
    """Scaffold a new experiment folder: subdirs + a prose README + a starter
    experiment.yml (the structured metadata), then index it."""
    parsed = _meta.parse_experiment_dirname(f"{args.exp_id} - {args.name}")
    if not parsed:
        die(f"invalid experiment id/name (expected K1-YYMMXX): {args.exp_id!r}")
    folder = store.home / f"{args.exp_id} - {args.name}"
    if folder.exists():
        die(f"folder already exists: {folder}")
    for sub in ("raw", "data", "protocol", "reports", "analysis"):
        (folder / sub).mkdir(parents=True, exist_ok=True)
    (folder / "README.md").write_text(
        _meta.readme_template({"exp_id": args.exp_id, "name": args.name}), encoding="utf-8")
    meta = provenance.validate({
        "exp_id": args.exp_id, "name": args.name, "status": "planned",
        "cro": args.cro, "model": args.model,
        "cro_study_ids": [args.study_id] if args.study_id else [],
    })
    provenance.write_sidecar(folder, meta)
    await _index_experiment(store, folder.resolve(), parsed, verbose=not args.json)
    if args.json:
        emit_json({"created": store.relpath(folder), "exp_id": args.exp_id})
    else:
        print(f"created {store.relpath(folder)} (raw/ data/ protocol/ reports/ analysis/ "
              f"+ README.md + experiment.yml)")
        print(f"indexed as {args.exp_id}; write README.md prose + fill experiment.yml")


def _parse_routes(raw: list[str] | None) -> dict[str, str]:
    """Parse repeatable ``--route "NAME=subdir"`` flags into a {name: subdir} map.

    NAME is a source file's basename; subdir must be one of LAYOUT's subfolders. The
    agent supplies these after reading the delivery — the *content* judgment intake no
    longer guesses (a document's role depends on what it contains)."""
    routes: dict[str, str] = {}
    for item in raw or []:
        name, sep, sub = item.partition("=")
        name, sub = name.strip(), sub.strip().lower()
        if not sep or not name or not sub:
            die(f"bad --route {item!r}: expected NAME=subdir "
                f"(subdir one of {', '.join(_intake.SUBDIRS)})")
        if sub not in _intake.SUBDIRS:
            die(f"bad --route {item!r}: subdir must be one of {', '.join(_intake.SUBDIRS)}")
        routes[name] = sub
    return routes


async def cmd_intake(store: Store, args: argparse.Namespace) -> None:
    """File a delivery (folder or files) into an experiment per LAYOUT.md.

    Copies (never moves) from the source; dry-run by default — review the plan,
    then re-run with --commit to copy + index. A document's *role* (protocol vs
    reports vs raw) is the agent's judgment, supplied per file with repeatable
    `--route "NAME=subdir"`; unrouted files fall back to a format/`raw` default the
    dry-run marks as a guess to confirm. See references/search-index.md.
    """
    import shutil

    src = Path(args.source).expanduser()
    if not src.exists():
        die(f"source not found: {src}")
    sources = sorted(p for p in src.rglob("*") if p.is_file()) if src.is_dir() else [src]

    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r} — scaffold it first with `sci new`")
    exp_dir, parsed = found
    routes = _parse_routes(getattr(args, "route", None))
    plan = _intake.plan_intake(sources, exp_dir, routes=routes)

    if args.json and not args.commit:
        emit_json({"experiment": parsed["exp_id"], "dry_run": True,
                   "plan": [{"src": str(p["src"]), "dest": store.relpath(p["dest"]),
                             "subdir": p["subdir"], "routed_by": p["routed_by"],
                             "collision": p["exists"]} for p in plan]})
        return
    if not args.commit:
        print(f"intake plan for {parsed['exp_id']} (dry-run — nothing copied):")
        by_sub: dict[str, int] = {}
        guessed = 0
        for p in plan:
            by_sub[p["subdir"]] = by_sub.get(p["subdir"], 0) + 1
            flag = "  ⚠ overwrites existing" if p["exists"] else ""
            mark = "  ? unreviewed default" if p["routed_by"] in ("default", "ext") else ""
            guessed += 1 if mark else 0
            print(f"  {p['subdir']:8} ← {p['src'].name}{flag}{mark}")
        print(f"  ({len(plan)} files: " + ", ".join(f"{n} {s}" for s, n in sorted(by_sub.items())) + ")")
        if guessed:
            print(f"  {guessed} file(s) fell back to a default placement — read them and re-route any "
                  f"protocol/reports/data with --route \"NAME=subdir\".")
        print("re-run with --commit to copy these in and index.")
        return

    copied = 0
    for p in plan:
        p["dest"].parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(p["src"], p["dest"])
        copied += 1
    result = await _index_experiment(store, exp_dir, parsed, verbose=not args.json)
    if args.json:
        emit_json({"experiment": parsed["exp_id"], "copied": copied, "indexed": result})
    else:
        print(f"copied {copied} files into {parsed['exp_id']} and reindexed "
              f"({result['files_indexed']} files total)")


async def cmd_catalog(store: Store, args: argparse.Namespace) -> None:
    """Export the experiment catalog: CATALOG.md (human index) + catalog.json."""
    import json

    exps = await store.experiments()
    exps.sort(key=lambda r: r.get("exp_id") or "")
    clean = [{k: v for k, v in e.items() if not k.startswith("_") and k != "content_hash"}
             for e in exps]
    md_path = store.home / "CATALOG.md"
    json_path = store.home / store._store_dirname / "catalog.json"
    md_path.write_text(_meta.catalog_markdown(exps), encoding="utf-8")
    json_path.write_text(json.dumps(clean, ensure_ascii=False, indent=2, default=str, sort_keys=True),
                         encoding="utf-8")
    if args.json:
        emit_json({"experiments": len(exps), "markdown": store.relpath(md_path),
                   "json": store.relpath(json_path)})
    else:
        print(f"wrote {store.relpath(md_path)} and {store.relpath(json_path)} "
              f"({len(exps)} experiments)")


async def _experiment_dirs(store: Store, only: str | None):
    """Yield (exp_dir, exp_id) for one experiment or all of them."""
    if only:
        found = _find_experiment_dir(store.home, only)
        if not found:
            die(f"no experiment matching {only!r}")
        yield found[0], found[1]["exp_id"]
        return
    for child in sorted(store.home.iterdir()):
        parsed = _meta.parse_experiment_dirname(child.name) if child.is_dir() else None
        if parsed:
            yield child.resolve(), parsed["exp_id"]


async def cmd_check(store: Store, args: argparse.Namespace) -> None:
    """Deterministic structural integrity report (reports only; never mutates)."""
    worklist = []
    async for exp_dir, exp_id in _experiment_dirs(store, args.experiment):
        rec = await store.get_experiment(exp_id) or {"exp_id": exp_id}
        files = await store.files(exp_id)
        flags = _audit.structural_flags(store.home, exp_dir, rec, files)
        if flags:
            worklist.append({"exp_id": exp_id, "flags": flags})
    if args.json:
        emit_json(worklist)
        return
    if not worklist:
        print("✓ no structural issues found")
        return
    for item in worklist:
        print(f"{item['exp_id']}:")
        for f in item["flags"]:
            print(f"    {f}")


def _staleness_entry(home: Path, exp_dir: Path, exp_id: str) -> dict[str, Any]:
    """The per-experiment provenance-staleness portion of an audit entry — PURE: it
    re-hashes the recorded ledger against the evidence on disk via the shared core, and
    never touches the libkit store. Shared by the store-backed ``cmd_audit`` and the
    store-free ``audit_report``."""
    sidecar_path = exp_dir / provenance.SIDECAR_NAME
    entry: dict[str, Any] = {"exp_id": exp_id}
    if not sidecar_path.is_file():
        entry["staleness"] = "no-experiment-yml"
        return entry
    try:
        provenance.read_sidecar(exp_dir)  # validate (raises on a bad sidecar)
        st = provenance.staleness(exp_dir, repo_root=home)
        entry["staleness"] = st["state"]
        if st["state"] == "stale":
            for k in ("changed", "missing", "added", "artifact_changed", "reviewed_at"):
                if st.get(k):
                    entry[k] = st[k]
    except provenance.SidecarError as e:
        entry["staleness"] = "invalid-experiment-yml"
        entry["error"] = str(e)
    return entry


def _source_files_on_disk(exp_dir: Path, home: Path) -> list[str]:
    """The data/report/raw/analysis files an agent should read for the semantic pass,
    as home-relative paths — derived by walking the folder, no store required."""
    out = []
    for f in _files.iter_experiment_files(exp_dir):
        if f["role"] in ("data", "report", "raw", "analysis"):
            try:
                out.append(str(f["abs_path"].resolve().relative_to(home.resolve())))
            except ValueError:
                out.append(f["filename"])
    return sorted(out)


def audit_report(home: Path, only: str | None = None) -> list[dict[str, Any]]:
    """Build the provenance-staleness audit report by walking on-disk experiment folders
    — NO libkit store required. One entry per experiment (or just ``only``), each with its
    ``staleness`` state (+ drift detail when stale) and a ``source_files`` worklist for
    the semantic pass. This is the store-free path `sci audit` uses when no store exists.
    """
    report = []
    if only:
        found = _find_experiment_dir(home, only)
        if not found:
            die(f"no experiment matching {only!r}")
        pairs = [(found[0], found[1]["exp_id"])]
    else:
        pairs = [(c.resolve(), p["exp_id"]) for c in sorted(home.iterdir())
                 if c.is_dir() and (p := _meta.parse_experiment_dirname(c.name))]
    for exp_dir, exp_id in pairs:
        entry = _staleness_entry(home, exp_dir, exp_id)
        entry["source_files"] = _source_files_on_disk(exp_dir, home)
        report.append(entry)
    return report


def print_audit_report(report: list[dict[str, Any]], as_json: bool) -> None:
    """Render an audit report (store-backed or store-free) — identical output either way."""
    if as_json:
        emit_json(report)
        return
    for e in report:
        print(f"{e['exp_id']}: {e['staleness']}")
        if e.get("error"):
            print(f"    {e['error']}")
        if e.get("staleness") == "stale":
            if e.get("artifact_changed"):
                print("    an artifact (e.g. README) edited since last review")
            for p in e.get("changed", []):
                print(f"    changed: {p}")
            for p in e.get("missing", []):
                print(f"    missing: {p}")
            for p in e.get("added", []):
                print(f"    added (unrecorded): {p}")
            print(f"    last reviewed {e.get('reviewed_at')}")
    print("\nFor the semantic pass, run `sci audit --json` and fan out an agent per "
          "experiment to read its source_files and verify the README/reports prose — "
          "including the prose ↔ claims check (every quantitative or qualitative result "
          "maps to a grounded claim, else flag it; see references/review-audit.md).")


async def cmd_audit(store: Store, args: argparse.Namespace) -> None:
    """Provenance staleness (experiment.yml provenance vs the evidence on disk) +
    a worklist for the parallel-agent semantic pass.

    Checks the WHOLE provenance ledger (data/ extract edges + analysis/ derive edges
    + the README review edge) via the shared core — re-hashing every recorded input
    and artifact and reporting per-file drift. When a store is present the
    ``source_files`` worklist is sourced from the index; the staleness check itself is
    store-free (see :func:`_staleness_entry`).
    """
    report = []
    async for exp_dir, exp_id in _experiment_dirs(store, args.experiment):
        files = await store.files(exp_id)
        entry = _staleness_entry(store.home, exp_dir, exp_id)
        # source files an agent should read to verify the prose semantically
        entry["source_files"] = [fr["path"] for fr in files
                                 if fr.get("role") in ("data", "report", "raw", "analysis")]
        report.append(entry)
    print_audit_report(report, args.json)


async def cmd_meta(store: Store, args: argparse.Namespace) -> None:
    """Show an experiment's structured metadata (from experiment.yml).

    Authoring the sidecar from a README is a reading task the agent does directly —
    see references/search-index.md ("Author experiment.yml from the README"); the
    tool never writes the sidecar from prose."""
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r}")
    exp_dir, parsed = found
    sidecar = exp_dir / provenance.SIDECAR_NAME

    if not sidecar.is_file():
        die(f"no {provenance.SIDECAR_NAME} for {parsed['exp_id']} — author one by reading "
            f"the README (references/search-index.md), or scaffold with `sci new`")
    try:
        meta = provenance.read_sidecar(exp_dir)
    except provenance.SidecarError as e:
        die(str(e))
    emit_json(meta)


async def cmd_fingerprint(store: Store, args: argparse.Namespace) -> None:
    """Show the input set `review` would record for an experiment's README right now —
    the in-folder data files (+ any externally-declared inputs), each with its current
    sha256. Lets you see exactly what provenance will track."""
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r}")
    exp_dir, parsed = found
    sidecar = exp_dir / provenance.SIDECAR_NAME
    declared = []
    if sidecar.is_file():
        meta = provenance.read_sidecar(exp_dir)
        exp_rel = exp_dir.resolve().relative_to(store.home.resolve()).as_posix()
        entry = provenance.provenance_entry(meta, provenance.DEFAULT_ARTIFACT) or {}
        declared = [i["path"] for i in (entry.get("inputs") or [])
                    if not i["path"].startswith(exp_rel + "/")]
    inputs, missing = provenance.resolve_inputs(store.home, exp_dir, declared)
    if args.json:
        emit_json({"exp_id": parsed["exp_id"], "inputs": inputs, "missing": missing})
        return
    for i in inputs:
        print(f"  {i['sha256']}  {i['path']}")
    print(f"({len(inputs)} input files" + (f", {len(missing)} declared-but-missing" if missing else "") + ")")
    for m in missing:
        print(f"  ! missing: {m}", file=sys.stderr)


async def cmd_review(store: Store, args: argparse.Namespace) -> None:
    """Mark an experiment's README as verified against its data: record an explicit
    input list (with each file's sha256) + the README's sha + a review date in
    experiment.yml. In-folder data files are included automatically; declare any
    external dependency (e.g. CRO slides under Shared/) with --input <repo-rel path>
    (repeatable; preserved across re-reviews). `audit` then reports per-file drift."""
    found = _find_experiment_dir(store.home, args.experiment)
    if not found:
        die(f"no experiment matching {args.experiment!r}")
    exp_dir, parsed = found
    sidecar = exp_dir / provenance.SIDECAR_NAME
    if not sidecar.is_file():
        die(f"no {provenance.SIDECAR_NAME} for {parsed['exp_id']} — create one first")
    try:
        meta = provenance.read_sidecar(exp_dir)
    except provenance.SidecarError as e:
        die(str(e))
    updated, missing = provenance.review(store.home, exp_dir, meta, today=args.date,
                                         extra_inputs=args.input or [])
    provenance.write_sidecar(exp_dir, updated)
    entry = provenance.provenance_entry(updated, provenance.DEFAULT_ARTIFACT)
    if entry is None:  # review() always writes this artifact entry; guard for robustness
        die(f"review did not record a {provenance.DEFAULT_ARTIFACT} provenance entry")
    if args.json:
        emit_json({"exp_id": parsed["exp_id"], "provenance": entry, "missing": missing})
    else:
        print(f"stamped {parsed['exp_id']}: README verified against "
              f"{len(entry.get('inputs') or [])} "
              f"input files (reviewed {entry.get('reviewed_at')})")
        for m in missing:
            print(f"  ! declared input not found on disk: {m}", file=sys.stderr)


async def cmd_pr(store: Store, args: argparse.Namespace) -> None:
    """Package working-tree changes into a branch + pull request for review."""
    await _maybe_pr(store, args.paths or None, args.title,
                    args.body or args.title, args, dry_run=args.dry_run)


async def _maybe_pr(store: Store, paths, title: str, body: str,
                    args: argparse.Namespace, *, dry_run: bool = False) -> None:
    try:
        result = _pr.create_pr(store.home, title=title, body=body,
                               paths=paths or _changed_paths(store.home),
                               dry_run=dry_run)
    except _pr.GitError as e:
        die(str(e))
    if args.json:
        emit_json(result)
    elif result.get("pr_url"):
        print(f"opened PR: {result['pr_url']}")
    elif result.get("dry_run"):
        print("dry-run — would run:\n  " + "\n  ".join(result["steps"]))
    else:
        print(f"committed to branch {result['branch']} (not pushed)")


def _changed_paths(home: Path) -> list[str]:
    import subprocess
    # -z gives NUL-separated, UNquoted paths (paths here contain spaces), so we can
    # hand them straight to `git add --` without quoting/escaping surprises.
    out = subprocess.run(["git", "-C", str(home), "status", "--porcelain", "-z"],
                         capture_output=True, text=True).stdout
    return [entry[3:] for entry in out.split("\0") if entry.strip()]


# --------------------------------------------------------------------------- #
# dispatch
# --------------------------------------------------------------------------- #
COMMANDS = {
    "init": cmd_init,
    "index": cmd_index,
    "reindex": cmd_reindex,
    "index-claims": cmd_index_claims,
    "list": cmd_list,
    "show": cmd_show,
    "search": cmd_search,
    "query": cmd_query,
    "file": cmd_file,
    "read": cmd_read,
    "entity": cmd_entity,
    "new": cmd_new,
    "intake": cmd_intake,
    "catalog": cmd_catalog,
    "check": cmd_check,
    "audit": cmd_audit,
    "meta": cmd_meta,
    "fingerprint": cmd_fingerprint,
    "review": cmd_review,
    "pr": cmd_pr,
}


def register(sub: argparse._SubParsersAction) -> None:
    """Register the store subcommands on an existing ``sci`` subparser action.

    Each store subcommand carries a ``--home`` flag (managed data folder; default
    ``$SCIENTIST_HOME``, else cwd) and a ``--json`` flag.
    """
    def add(name: str, help_: str) -> argparse.ArgumentParser:
        sp = sub.add_parser(name, help=help_)
        sp.add_argument("--home",
                        help="managed data folder (default: $SCIENTIST_HOME or cwd)")
        sp.add_argument("--json", action="store_true", help="machine-readable output")
        return sp

    add("init", "create the libkit store and .gitignore under the data folder")
    sp = add("index", "index one experiment folder (by exp_id or path)")
    sp.add_argument("experiment")
    add("reindex", "index every experiment folder under the data folder")
    sp = add("index-claims", "index grounded claims from an experiment's grounding_report.json")
    sp.add_argument("experiment")
    sp.add_argument("--report", help="grounding_report.json to index "
                    "(default <exp>/analysis/grounding_report.json then <exp>/grounding_report.json)")
    sp = add("list", "list experiments (default), files, entities, claims, or reports")
    sp.add_argument("--kind", choices=["experiment", "file", "entity", "claim", "report"],
                    default="experiment")
    sp.add_argument("--experiment", help="when --kind file/claim/report: limit to this exp_id")
    sp = add("show", "show one experiment and its files")
    sp.add_argument("experiment")
    sp = add("search", "metadata search across experiments and files")
    sp.add_argument("text")
    sp = add("query", "semantic + full-text search inside indexed content")
    sp.add_argument("text")
    sp.add_argument("--limit", type=int, default=8)
    sp.add_argument("--kind", choices=["experiment", "file", "entity", "claim", "report"],
                    default=None)
    sp = add("file", "show one file record (by relative path)")
    sp.add_argument("path")
    sp = add("read", "dump a tabular file (csv/tsv/xlsx) to stdout")
    sp.add_argument("path")
    sp = add("entity", "list derived entities or show one entity's experiments")
    sp.add_argument("entity_action", choices=["list", "show"])
    sp.add_argument("name", nargs="?", help="entity name (for show)")
    sp = add("new", "scaffold a new experiment folder (subdirs + README template)")
    sp.add_argument("exp_id", help="internal id, e.g. K1-000003")
    sp.add_argument("name", help="short name, e.g. 'Rat IT Chronic Tox'")
    sp.add_argument("--cro", help="contract research org")
    sp.add_argument("--study-id", help="external/CRO study id")
    sp.add_argument("--model", help="species/model")
    sp = add("intake", "file a delivery (folder/files) into an experiment per LAYOUT.md")
    sp.add_argument("experiment", help="target experiment (exp_id or folder)")
    sp.add_argument("source", help="a delivery folder or a single file (copied, not moved)")
    sp.add_argument("--route", action="append", metavar="NAME=SUBDIR",
                    help="place file NAME in SUBDIR (protocol/reports/data/raw/analysis); "
                         "repeatable — the agent's per-document role call")
    sp.add_argument("--commit", action="store_true", help="actually copy + index (default: dry-run)")
    add("catalog", "export the experiment catalog (CATALOG.md + catalog.json)")
    sp = add("check", "structural integrity report (missing/unindexed files, layout, redundant archives)")
    sp.add_argument("experiment", nargs="?", help="limit to one experiment (default: all)")
    sp = add("audit", "provenance staleness of the experiment.yml ledger + a worklist for the "
             "semantic pass (which includes the prose↔claims check)")
    sp.add_argument("experiment", nargs="?", help="limit to one experiment (default: all)")
    sp = add("meta", "show an experiment's structured metadata (experiment.yml)")
    sp.add_argument("experiment")
    sp = add("fingerprint", "show the input files (+ sha256) review would record for an experiment")
    sp.add_argument("experiment")
    sp = add("review", "record provenance after verifying the README vs its data (explicit input list)")
    sp.add_argument("experiment")
    sp.add_argument("--input", action="append", metavar="REPO_REL_PATH",
                    help="declare an external dependency (repeatable; e.g. a Shared/ CRO file)")
    sp.add_argument("--date", help="review date YYYY-MM-DD (default: today)")
    sp = add("pr", "package working-tree changes into a branch + pull request")
    sp.add_argument("title")
    sp.add_argument("paths", nargs="*", help="paths to include (default: all changes)")
    sp.add_argument("--body", help="PR body")
    sp.add_argument("--dry-run", action="store_true", help="show the git/gh steps, do nothing")


class _HomeOnly:
    """Lightweight stand-in for commands that need only the data folder, not the
    libkit store (so `pr` doesn't require an embedding backend)."""

    def __init__(self, home: Path) -> None:
        self.home = home

    def relpath(self, path: Path) -> str:
        try:
            return str(path.resolve().relative_to(self.home.resolve()))
        except ValueError:
            return str(path)


async def _run(args: argparse.Namespace) -> None:
    home = _home(args)
    _load_dotenv(home)
    handler = COMMANDS[args.cmd]
    if args.cmd == "init":
        home.mkdir(parents=True, exist_ok=True)
    else:
        _require_initialized(home)
    if args.cmd == "pr":            # pure git; no libkit store needed
        await handler(_HomeOnly(home), args)  # type: ignore[arg-type]
        return
    try:
        store = Store.open(home)
    except EmbedderConfigError as e:
        die(str(e))
    try:
        await handler(store, args)
    finally:
        await store.close()


def store_exists(args: argparse.Namespace) -> bool:
    """Whether a libkit store is initialized under the resolved data folder."""
    home = _home(args)
    return (home / STORE_DIRNAME / "catalog.duckdb").exists()


def dispatch_audit_storeless(args: argparse.Namespace) -> int:
    """Run `sci audit` WITHOUT opening the libkit store: pure provenance staleness over
    on-disk experiment folders. Used when no store is initialized so a single on-disk
    experiment can be audited without a store (and without a "no scientist store" error).
    """
    home = _home(args)
    _load_dotenv(home)
    report = audit_report(home, getattr(args, "experiment", None))
    print_audit_report(report, args.json)
    return 0


async def _run_index_report(args: argparse.Namespace, card: dict[str, Any]) -> None:
    home = _home(args)
    _load_dotenv(home)
    _require_initialized(home)
    try:
        store = Store.open(home)
    except EmbedderConfigError as e:
        die(str(e))
    try:
        rec = await store.upsert_report(card)
        if args.json:
            emit_json({"report_id": rec.get("report_id"), "scope": rec.get("scope"),
                       "exp_id": rec.get("exp_id"), "document_id": rec.get("document_id")})
        else:
            print(f"indexed report {rec.get('report_id')} "
                  f"({rec.get('scope')}{', ' + rec['exp_id'] if rec.get('exp_id') else ''}) "
                  f"into the store")
    finally:
        await store.close()


def index_report(args: argparse.Namespace, card: dict[str, Any]) -> int:
    """Open the libkit store and upsert a ``kind=report`` document from a prepared ``card``
    dict (built store-free by ``provenance.report`` + ``sci report``). Sync wrapper."""
    try:
        asyncio.run(_run_index_report(args, card))
    except KeyboardInterrupt:
        die("interrupted", code=130)
    return 0


def dispatch(args: argparse.Namespace) -> int:
    """Run the selected store subcommand (sync wrapper over the async handler)."""
    try:
        asyncio.run(_run(args))
    except KeyboardInterrupt:
        die("interrupted", code=130)
    return 0
