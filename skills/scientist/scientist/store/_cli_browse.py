"""Read-only browse/search command handlers: ``list``, ``show``, ``search``,
``query``, ``file``, ``read``, ``entity``.

These never mutate the store (they are the ``_READ_ONLY_COMMANDS``): metadata
listing, substring search, libkit hybrid query, single-record lookup, a
format-aware tabular dump, and the derived-entity views.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from ..cli_utils import die, emit_json
from . import _generate, _meta
from ._store import Store


async def cmd_list(store: Store, args: argparse.Namespace) -> None:
    if args.kind == "file":
        recs = await store.files(args.experiment)
    elif args.kind == "entity":
        recs = await store.all_records({"kind": "entity"})
    elif args.kind == "claim":
        recs = await store.claims(args.experiment)
    elif args.kind == "report":
        recs = await store.reports(args.experiment)
    elif args.kind == "litreview":
        recs = await store.litreviews()
    else:
        recs = await store.experiments()
    recs.sort(key=lambda r: r.get("exp_id") or r.get("path")
              or r.get("entity_id") or r.get("claim_id") or r.get("report_id")
              or r.get("litreview_id") or "")
    if args.json:
        emit_json(recs)
        return
    if not recs:
        print("(nothing indexed)")
        return
    for r in recs:
        if args.kind == "file":
            print(f"  [{r.get('role','?'):8}] {r.get('path')}  ({r.get('indexed_as','?')})")
        elif args.kind == "entity":
            print(f"  {r.get('entity_id')}  — {r.get('title') or ''}")
        elif args.kind == "claim":
            outcome = r.get("outcome") or "?"
            label = _meta.CLAIM_OUTCOME_LABEL.get(outcome, outcome)
            stmt = (r.get("statement") or "").strip().replace("\n", " ")[:90]
            print(f"  [{label} · {r.get('strength','?')}] {r.get('exp_id')}  {stmt}")
        elif args.kind == "report":
            title = (r.get("title") or r.get("slug") or "").strip().replace("\n", " ")[:80]
            print(f"  [{r.get('scope','?')}] {r.get('report_id')}  {title}")
        elif args.kind == "litreview":
            title = (r.get("title") or r.get("slug") or "").strip().replace("\n", " ")[:80]
            n = len(r.get("must_confront") or [])
            print(f"  [{r.get('scope','?')}] {r.get('litreview_id')}  {title}  ({n} must-confront)")
        else:
            fc = r.get("file_counts") or {}
            print(f"  {r.get('exp_id')}  {r.get('name') or r.get('title') or ''}"
                  f"   ({fc.get('files_indexed', 0)} files)")


async def cmd_show(store: Store, args: argparse.Namespace) -> None:
    rec = await store.get_experiment(args.experiment)
    if rec is None:
        die(f"no experiment {args.experiment!r} (index it with `sci index`)")
    files = await store.files(args.experiment)
    if args.json:
        emit_json({"experiment": rec, "files": files})
        return
    print(f"{rec.get('exp_id')}: {rec.get('title') or rec.get('name')}")
    for label, key in (("CRO study IDs", "cro_study_ids"), ("CRO", "cro"),
                       ("Status", "status"), ("Folder", "folder")):
        v = rec.get(key)
        if v:
            print(f"  {label}: {', '.join(v) if isinstance(v, list) else v}")
    for label, key in (("Assays", "assays"), ("ASOs", "asos"), ("Model", "model")):
        v = rec.get(key)
        if v:
            print(f"  {label}: {', '.join(v) if isinstance(v, list) else v}")
    print(f"\n  Files ({len(files)}):")
    for line in _generate.files_on_disk_table(files).splitlines():
        print(f"  {line}")


async def cmd_search(store: Store, args: argparse.Namespace) -> None:
    """Metadata search across experiments + files (substring over key fields)."""
    needle = args.text.lower()
    hits = []
    for r in await store.all_records():
        hay = " ".join(str(r.get(k, "")) for k in
                       ("exp_id", "name", "title", "cro", "path", "role", "filename")).lower()
        hay += " ".join(str(x) for x in (r.get("cro_study_ids") or []) +
                        (r.get("assays") or []) + (r.get("asos") or []) + (r.get("tags") or [])).lower()
        if needle in hay:
            hits.append(r)
    if args.json:
        emit_json(hits)
        return
    if not hits:
        print("(no matches)")
        return
    for r in hits:
        if r.get("kind") == "experiment":
            print(f"  [exp]  {r.get('exp_id')}  {r.get('title') or r.get('name')}")
        else:
            print(f"  [file] {r.get('exp_id')}  {r.get('path')}")


async def cmd_query(store: Store, args: argparse.Namespace) -> None:
    """Semantic + full-text search inside indexed content (libkit hybrid)."""
    filters = {"kind": args.kind} if args.kind else None
    results = await store.query(args.text, limit=args.limit, filters=filters)
    out = []
    for r in results:
        chunk = r.chunk
        meta = chunk.metadata or {}
        hit = {
            "score": r.score,
            "exp_id": meta.get("exp_id"),
            "path": meta.get("path"),
            "kind": meta.get("kind"),
            "text": chunk.text,
        }
        if meta.get("kind") == "claim":
            # Surface the judgment so a contradicted/weak claim is never shown as
            # plain positive evidence.
            hit["outcome"] = meta.get("outcome")
            hit["strength"] = meta.get("strength")
            hit["claim_kind"] = meta.get("claim_kind")
            hit["statement"] = meta.get("statement")
            hit["claim_id"] = meta.get("claim_id")
        elif meta.get("kind") == "report":
            hit["report_id"] = meta.get("report_id")
            hit["scope"] = meta.get("scope")
            hit["report_title"] = meta.get("title")
        elif meta.get("kind") == "litreview":
            hit["litreview_id"] = meta.get("litreview_id")
            hit["scope"] = meta.get("scope")
            hit["litreview_title"] = meta.get("title")
        out.append(hit)
    if args.json:
        emit_json(out)
        return
    if not out:
        print("(no results)")
        return
    for h in out:
        if h.get("kind") == "claim":
            outcome = h.get("outcome") or "?"
            label = _meta.CLAIM_OUTCOME_LABEL.get(outcome, outcome)
            stmt = (h.get("statement") or h.get("text") or "").strip().replace("\n", " ")[:200]
            print(f"  [claim · {label} · strength: {h.get('strength','?')}] {h.get('exp_id')}\n"
                  f"      {stmt}")
            continue
        if h.get("kind") == "report":
            snippet = (h.get("text") or "").strip().replace("\n", " ")[:200]
            print(f"  [report · {h.get('scope','?')}] {h.get('report_id')}"
                  f"  {h.get('report_title') or ''}\n      {snippet}")
            continue
        if h.get("kind") == "litreview":
            snippet = (h.get("text") or "").strip().replace("\n", " ")[:200]
            print(f"  [litreview · {h.get('scope','?')}] {h.get('litreview_id')}"
                  f"  {h.get('litreview_title') or ''}\n      {snippet}")
            continue
        loc = h.get("path") or h.get("exp_id") or "?"
        snippet = (h.get("text") or "").strip().replace("\n", " ")[:200]
        print(f"  {loc}\n      {snippet}")


async def cmd_file(store: Store, args: argparse.Namespace) -> None:
    rec = await store.get_file(args.path)
    if rec is None:
        die(f"no file record for {args.path!r}")
    emit_json(rec)


async def cmd_read(store: Store, args: argparse.Namespace) -> None:
    """Format-aware dump of a tabular file to stdout (for pulling exact values)."""
    path = (store.home / args.path) if not Path(args.path).is_absolute() else Path(args.path)
    if not path.exists():
        die(f"file not found: {path}")
    ext = path.suffix.lower()
    if ext in (".csv", ".tsv", ".xlsx", ".xlsm"):
        if ext in (".csv", ".tsv"):
            print(path.read_text(encoding="utf-8", errors="replace"))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    print(f"# sheet: {ws.title}")
                    for row in ws.iter_rows(values_only=True):
                        print("\t".join("" if c is None else str(c) for c in row))
            finally:
                wb.close()
    else:
        die(f"`read` handles csv/tsv/xlsx; {ext or 'this file'} should be opened directly "
            f"(path: {path})")


async def cmd_entity(store: Store, args: argparse.Namespace) -> None:
    """Entities are derived live from experiment records (registry), plus any
    curated notes (kind=entity). `list` aggregates; `show` filters experiments."""
    exps = await store.experiments()
    if args.entity_action == "list":
        agg: dict[str, dict[str, set]] = {"asos": {}, "assays": {}, "cro": {}}
        for e in exps:
            for fld, key in (("asos", "asos"), ("assays", "assays")):
                for v in e.get(key) or []:
                    agg["asos" if fld == "asos" else "assays"].setdefault(v, set()).add(e["exp_id"])
            if e.get("cro"):
                agg["cro"].setdefault(e["cro"], set()).add(e["exp_id"])
        if args.json:
            emit_json({k: {name: sorted(ids) for name, ids in d.items()} for k, d in agg.items()})
            return
        for kind, d in agg.items():
            if d:
                print(f"{kind}:")
                for name, ids in sorted(d.items()):
                    print(f"  {name}  ({len(ids)} experiments)")
    else:  # show
        ident = args.name
        matched = [e["exp_id"] for e in exps
                   if ident in (e.get("asos") or []) or ident in (e.get("assays") or [])
                   or ident == e.get("cro") or ident in (e.get("cro_study_ids") or [])]
        note = await store.get_entity(_meta.entity_slug(ident))
        out = {"entity": ident, "experiments": sorted(matched),
               "curated_note": (note or {}).get("note")}
        emit_json(out) if args.json else print(
            f"{ident}: {len(matched)} experiments\n  " + "\n  ".join(sorted(matched)))
